import openpyxl

wb = openpyxl.load_workbook("nilai cpl mhs/Pengukuran CPL SIA SMI Genap 2024-2025.xlsx", data_only=True)

sheets = ['CPL PRODI', 'Pengukuran CPL02', 'Pengukuran CPL03', 'Pengukuran CPL04', 'PENGUKURAN CPL ALL']
for name in sheets:
    if name in wb.sheetnames:
        print(f"\n--- Sheet: {name} ---")
        sheet = wb[name]
        rows = list(sheet.iter_rows(max_row=10, max_col=10, values_only=True))
        for i, r in enumerate(rows):
            print(f"  Row {i+1}: {r}")
